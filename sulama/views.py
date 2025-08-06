from django.shortcuts import render
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Count, Sum, Avg
from django.db.models.functions import Extract
from datetime import datetime, timedelta
from authentication.permissions import SulamaYetkisiPermission
from authentication.mixins import SulamaBazliMixin
import shutil
from django.http import HttpResponse, JsonResponse
from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
import json
import os
from .models import (
    Bolge, Sulama, DepolamaTesisi, Kanal, 
    GunlukSebekeyeAlinanSuMiktari, GunlukDepolamaTesisiSuMiktari,
    UrunKategorisi, Urun, YillikGenelSuTuketimi, YillikUrunDetay,
    Makina, MakinaKonum, MakinaIs
)
from .serializers import (
    BolgeSerializer, SulamaSerializer, DepolamaTesisiSerializer, KanalSerializer,
    GunlukSebekeyeAlinanSuMiktariSerializer, GunlukDepolamaTesisiSuMiktariSerializer,
    UrunKategorisiSerializer, UrunSerializer, 
    YillikGenelSuTuketimiSerializer, YillikUrunDetaySerializer, 
    SulamaOzetSerializer, KanalOzetSerializer, UrunOzetSerializer,
    MakinaSerializer, MakinaKonumSerializer, MakinaIsSerializer
)
from django.utils import timezone


class BolgeViewSet(viewsets.ModelViewSet):
    """
    Bölge yönetimi ViewSet
    """
    queryset = Bolge.objects.all()
    serializer_class = BolgeSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['isim']
    search_fields = ['isim', 'aciklama', 'yonetici']
    ordering_fields = ['isim', 'olusturma_tarihi']
    ordering = ['isim']


class SulamaViewSet(SulamaBazliMixin, viewsets.ModelViewSet):
    """
    Sulama sistemi yönetimi ViewSet
    Kullanıcı sadece yetkili olduğu sulama sistemlerini görebilir/yönetebilir
    """
    queryset = Sulama.objects.select_related('bolge').all()
    serializer_class = SulamaSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['bolge', 'isim']
    search_fields = ['isim', 'aciklama', 'bolge__isim']
    ordering_fields = ['isim', 'olusturma_tarihi', 'bolge__isim']
    ordering = ['bolge__isim', 'isim']
    pagination_class = None  # Pagination'ı kaldır

    def get_queryset(self):
        """Kullanıcının yetkili olduğu sulama sistemlerini getir"""
        base_queryset = super().get_queryset()
        return self.filter_by_sulama_permission(base_queryset, 'id')

    @action(detail=False, methods=['get'])
    def ozet(self, request):
        """Sulama sistemleri özet listesi"""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = SulamaOzetSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def istatistikler(self, request, pk=None):
        """Sulama sistemi istatistikleri"""
        sulama = self.get_object()
        
        istatistikler = {
            'depolama_tesisi_sayisi': sulama.depolama_tesisleri.count(),
            'kanal_sayisi': Kanal.objects.filter(depolama_tesisi__sulama=sulama).count(),
            'urun_sayisi': sulama.urunler.count(),
            'toplam_yillik_alan': YillikGenelSuTuketimi.objects.filter(sulama=sulama).aggregate(
                toplam=Sum('alan'))['toplam'] or 0,
            'toplam_yillik_tuketim': YillikGenelSuTuketimi.objects.filter(sulama=sulama).aggregate(
                toplam=Sum('su_tuketimi'))['toplam'] or 0,
        }
        
        # Son 7 günlük veri sayısı
        son_hafta = datetime.now().date() - timedelta(days=7)
        istatistikler['son_hafta_veri_sayisi'] = GunlukSebekeyeAlinanSuMiktari.objects.filter(
            kanal__depolama_tesisi__sulama=sulama, 
            tarih__gte=son_hafta
        ).count()
        
        return Response(istatistikler)


class DepolamaTesisiViewSet(SulamaBazliMixin, viewsets.ModelViewSet):
    """
    Depolama tesisi ViewSet
    """
    queryset = DepolamaTesisi.objects.select_related('sulama__bolge').all()
    serializer_class = DepolamaTesisiSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['sulama', 'isim']
    search_fields = ['isim', 'aciklama', 'sulama__isim']
    ordering_fields = ['isim', 'olusturma_tarihi']
    ordering = ['sulama', 'isim']
    pagination_class = None  # Pagination'ı kaldır

    def get_queryset(self):
        """Kullanıcının yetkili olduğu sulama sistemlerine ait tesisleri getir"""
        base_queryset = super().get_queryset()
        return self.filter_by_sulama_permission(base_queryset, 'sulama')

    @action(detail=True, methods=['post'])
    def su_hacmi_hesapla(self, request, pk=None):
        """Depolama tesisi ID ve kot değerine göre su hacmini hesapla"""
        depolama_tesisi = self.get_object()
        kot = request.data.get('kot')
        
        if kot is None:
            return Response(
                {'error': 'Kot parametresi gerekli'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            kot = float(kot)
            
            # Depolama tesisi abağından su miktarını hesapla
            try:
                from .models import DepolamaTesisiAbak
                abak = depolama_tesisi.abaklar.get(kot=kot)
                return Response({
                    'depolama_tesisi_id': depolama_tesisi.id,
                    'depolama_tesisi_isim': depolama_tesisi.isim,
                    'kot': kot,
                    'su_hacmi': abak.hacim,
                    'success': True
                })
            except DepolamaTesisiAbak.DoesNotExist:
                return Response({
                    'error': f'Bu depolama tesisi için {kot} m kot değeri abakta bulunamadı',
                    'depolama_tesisi_id': depolama_tesisi.id,
                    'depolama_tesisi_isim': depolama_tesisi.isim,
                    'kot': kot,
                    'su_hacmi': 0,
                    'success': False
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except ValueError:
            return Response(
                {'error': 'Geçersiz kot değeri'}, 
                status=status.HTTP_400_BAD_REQUEST
            )


class KanalViewSet(SulamaBazliMixin, viewsets.ModelViewSet):
    """
    Kanal yönetimi ViewSet
    """
    queryset = Kanal.objects.select_related('depolama_tesisi__sulama__bolge').all()
    serializer_class = KanalSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['depolama_tesisi', 'isim', 'kanal_kodu']
    search_fields = ['isim', 'aciklama', 'kanal_kodu', 'depolama_tesisi__isim']
    ordering_fields = ['isim', 'olusturma_tarihi', 'kanal_kodu']
    ordering = ['depolama_tesisi', 'isim']
    pagination_class = None  # Pagination'ı kaldır

    def get_queryset(self):
        """Kullanıcının yetkili olduğu sulama sistemlerine ait kanalları getir"""
        base_queryset = super().get_queryset()
        # Sulama bazlı filtreleme
        return self.filter_by_sulama_permission(base_queryset, 'depolama_tesisi__sulama')

    @action(detail=False, methods=['get'])
    def ozet(self, request):
        """Kanal özet listesi"""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = KanalOzetSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def son_veriler(self, request, pk=None):
        """Kanalın son veri kayıtları"""
        kanal = self.get_object()
        son_veriler = GunlukSebekeyeAlinanSuMiktari.objects.filter(
            kanal=kanal
        ).order_by('-tarih', '-baslangic_saati')[:10]
        
        serializer = GunlukSebekeyeAlinanSuMiktariSerializer(son_veriler, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def su_hacmi_hesapla(self, request, pk=None):
        """Kanal ID ve yükseklik değerine göre su hacmini hesapla"""
        kanal = self.get_object()
        yukseklik = request.data.get('yukseklik')
        
        if yukseklik is None:
            return Response(
                {'error': 'Yükseklik parametresi gerekli'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            yukseklik = float(yukseklik)
            
            # Kanal abağından su miktarını hesapla
            try:
                from .models import KanalAbak
                abak = kanal.abaklar.get(yukseklik=yukseklik)
                return Response({
                    'kanal_id': kanal.id,
                    'kanal_isim': kanal.isim,
                    'yukseklik': yukseklik,
                    'su_hacmi': abak.hacim,
                    'success': True
                })
            except KanalAbak.DoesNotExist:
                return Response({
                    'error': f'Bu kanal için {yukseklik} m yükseklik değeri abakta bulunamadı',
                    'kanal_id': kanal.id,
                    'kanal_isim': kanal.isim,
                    'yukseklik': yukseklik,
                    'su_hacmi': 0,
                    'success': False
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except ValueError:
            return Response(
                {'error': 'Geçersiz yükseklik değeri'}, 
                status=status.HTTP_400_BAD_REQUEST
            )


class GunlukSebekeyeAlinanSuMiktariViewSet(SulamaBazliMixin, viewsets.ModelViewSet):
    """
    Günlük şebekeye alınan su miktarı ViewSet
    """
    queryset = GunlukSebekeyeAlinanSuMiktari.objects.select_related(
        'kanal__depolama_tesisi__sulama__bolge'
    ).all()
    serializer_class = GunlukSebekeyeAlinanSuMiktariSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['kanal', 'tarih', 'kanal__depolama_tesisi']
    search_fields = ['kanal__isim', 'kanal__depolama_tesisi__isim']
    ordering_fields = ['tarih', 'baslangic_saati', 'su_miktari']
    ordering = ['-tarih', '-baslangic_saati']
    pagination_class = None  # Pagination'ı kaldır

    def get_queryset(self):
        """Kullanıcının yetkili olduğu sulama sistemlerine ait verileri getir"""
        base_queryset = super().get_queryset()
        filtered_queryset = self.filter_by_sulama_permission(base_queryset, 'kanal__depolama_tesisi__sulama')
        
        # URL parametrelerinden tarih filtreleme
        baslangic_tarih = self.request.query_params.get('baslangic_tarih')
        bitis_tarih = self.request.query_params.get('bitis_tarih')
        
        if baslangic_tarih and bitis_tarih:
            try:
                baslangic = datetime.strptime(baslangic_tarih, '%Y-%m-%d').date()
                bitis = datetime.strptime(bitis_tarih, '%Y-%m-%d').date()
                filtered_queryset = filtered_queryset.filter(tarih__gte=baslangic, tarih__lte=bitis)
            except ValueError:
                pass  # Geçersiz tarih formatı durumunda varsayılan filtreleme
        
        # Eğer hiç tarih filtresi yoksa, mevcut ayın verilerini getir
        if not baslangic_tarih and not bitis_tarih:
            from datetime import date
            today = date.today()
            # Ayın ilk günü
            first_day = today.replace(day=1)
            # Bir sonraki ayın ilk günü
            if today.month == 12:
                next_month = today.replace(year=today.year + 1, month=1, day=1)
            else:
                next_month = today.replace(month=today.month + 1, day=1)
            
            filtered_queryset = filtered_queryset.filter(tarih__gte=first_day, tarih__lt=next_month)
        
        return filtered_queryset

    @action(detail=False, methods=['get'])
    def tarih_araligi(self, request):
        """Belirli tarih aralığındaki veriler"""
        baslangic = request.query_params.get('baslangic')
        bitis = request.query_params.get('bitis')
        
        if not baslangic or not bitis:
            return Response(
                {'error': 'Başlangıç ve bitiş tarihi gerekli'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            baslangic_tarih = datetime.strptime(baslangic, '%Y-%m-%d').date()
            bitis_tarih = datetime.strptime(bitis, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {'error': 'Geçersiz tarih formatı (YYYY-MM-DD kullanın)'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.get_queryset().filter(
            tarih__gte=baslangic_tarih, 
            tarih__lte=bitis_tarih
        )
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def ozet_istatistik(self, request):
        """Su miktarı özet istatistikleri"""
        queryset = self.filter_queryset(self.get_queryset())
        
        istatistikler = queryset.aggregate(
            toplam_su=Sum('su_miktari'),
            ortalama_su=Avg('su_miktari'),
            kayit_sayisi=Count('id')
        )
        
        return Response(istatistikler)

    @action(detail=False, methods=['post'])
    def hesapla_su_miktari(self, request):
        """Yükseklik değerine göre su miktarını hesapla"""
        kanal_id = request.data.get('kanal')
        yukseklik = request.data.get('yukseklik')
        
        if not kanal_id or yukseklik is None:
            return Response(
                {'error': 'Kanal ve yükseklik parametreleri gerekli'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            yukseklik = float(yukseklik)
            kanal = Kanal.objects.get(id=kanal_id)
            
            # Yetki kontrolü
            if not self.check_sulama_permission(kanal.depolama_tesisi.sulama.id):
                return Response(
                    {'error': 'Bu kanala erişim yetkiniz yok'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Kanal abağından su miktarını hesapla
            try:
                from .models import KanalAbak
                abak = kanal.abaklar.get(yukseklik=yukseklik)
                return Response({
                    'su_miktari': abak.hacim,
                    'yukseklik': yukseklik,
                    'kanal': kanal.isim,
                    'success': True
                })
            except KanalAbak.DoesNotExist:
                return Response({
                    'error': f'Bu kanal için {yukseklik} m yükseklik değeri abakta bulunamadı',
                    'su_miktari': 0,
                    'success': False
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except (ValueError, Kanal.DoesNotExist) as e:
            return Response(
                {'error': 'Geçersiz kanal veya yükseklik değeri'}, 
                status=status.HTTP_400_BAD_REQUEST
            )


class GunlukDepolamaTesisiSuMiktariViewSet(SulamaBazliMixin, viewsets.ModelViewSet):
    """
    Günlük depolama tesisi su miktarı ViewSet
    """
    queryset = GunlukDepolamaTesisiSuMiktari.objects.select_related(
        'depolama_tesisi__sulama__bolge'
    ).all()
    serializer_class = GunlukDepolamaTesisiSuMiktariSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['depolama_tesisi', 'tarih']
    search_fields = ['depolama_tesisi__isim', 'depolama_tesisi__sulama__isim']
    ordering_fields = ['tarih', 'su_miktari', 'kot']
    ordering = ['-tarih', 'depolama_tesisi']
    pagination_class = None  # Pagination'ı kaldır

    def get_queryset(self):
        """Kullanıcının yetkili olduğu sulama sistemlerine ait verileri getir"""
        base_queryset = super().get_queryset()
        filtered_queryset = self.filter_by_sulama_permission(base_queryset, 'depolama_tesisi__sulama')
        
        # URL parametrelerinden tarih filtreleme
        baslangic_tarih = self.request.query_params.get('baslangic_tarih')
        bitis_tarih = self.request.query_params.get('bitis_tarih')
        
        if baslangic_tarih and bitis_tarih:
            try:
                baslangic = datetime.strptime(baslangic_tarih, '%Y-%m-%d').date()
                bitis = datetime.strptime(bitis_tarih, '%Y-%m-%d').date()
                filtered_queryset = filtered_queryset.filter(tarih__gte=baslangic, tarih__lte=bitis)
            except ValueError:
                pass  # Geçersiz tarih formatı durumunda varsayılan filtreleme
        
        # Eğer hiç tarih filtresi yoksa, mevcut ayın verilerini getir
        if not baslangic_tarih and not bitis_tarih:
            from datetime import date
            today = date.today()
            # Ayın ilk günü
            first_day = today.replace(day=1)
            # Bir sonraki ayın ilk günü
            if today.month == 12:
                next_month = today.replace(year=today.year + 1, month=1, day=1)
            else:
                next_month = today.replace(month=today.month + 1, day=1)
            
            filtered_queryset = filtered_queryset.filter(tarih__gte=first_day, tarih__lt=next_month)
        
        return filtered_queryset

    @action(detail=False, methods=['get'])
    def son_durum(self, request):
        """Depolama tesislerinin son durumu"""
        # Her tesis için en son tarihdeki veriyi getir
        from django.db.models import Max
        
        queryset = self.get_queryset()
        son_kayitlar = []
        
        tesisler = DepolamaTesisi.objects.filter(
            gunluk_depolama_tesisi_su_miktarlari__in=queryset
        ).distinct()
        
        for tesis in tesisler:
            son_kayit = queryset.filter(depolama_tesisi=tesis).order_by('-tarih').first()
            if son_kayit:
                son_kayitlar.append(son_kayit)
        
        serializer = self.get_serializer(son_kayitlar, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def istatistikler(self, request):
        """Depolama tesisi su miktarı istatistikleri - Her tesisin son durumu"""
        from django.db.models import Max
        
        queryset = self.get_queryset()
        
        # URL parametrelerinden tarih filtreleme al
        baslangic_tarih = self.request.query_params.get('baslangic_tarih')
        bitis_tarih = self.request.query_params.get('bitis_tarih')
        depolama_tesisi = self.request.query_params.get('depolama_tesisi')
        
        # Filtreleri uygula
        if baslangic_tarih and bitis_tarih:
            try:
                baslangic = datetime.strptime(baslangic_tarih, '%Y-%m-%d').date()
                bitis = datetime.strptime(bitis_tarih, '%Y-%m-%d').date()
                queryset = queryset.filter(tarih__gte=baslangic, tarih__lte=bitis)
            except ValueError:
                pass
        
        if depolama_tesisi:
            queryset = queryset.filter(depolama_tesisi=depolama_tesisi)
        
        # Belirtilen tarih aralığında her depolama tesisinin en son kaydını al
        tesisler = DepolamaTesisi.objects.filter(
            id__in=queryset.values_list('depolama_tesisi', flat=True).distinct()
        )
        
        son_kayitlar = []
        for tesis in tesisler:
            # Bu tesis için belirtilen tarih aralığındaki en son kayıt
            son_kayit = queryset.filter(depolama_tesisi=tesis).order_by('-tarih').first()
            if son_kayit:
                son_kayitlar.append(son_kayit)
        
        if not son_kayitlar:
            return Response({
                'toplam_kayit': 0,
                'toplam_su_miktari': 0.0,
                'ortalama_su_miktari': 0.0,
                'aktif_tesis_sayisi': 0
            })
        
        # İstatistikleri hesapla
        toplam_su = sum(kayit.su_miktari for kayit in son_kayitlar)
        aktif_tesis_sayisi = len(son_kayitlar)
        ortalama_su = toplam_su / aktif_tesis_sayisi if aktif_tesis_sayisi > 0 else 0
        
        # Toplam kayıt sayısı da filtrelenmiş queryset'ten
        toplam_kayit = queryset.count()
        
        return Response({
            'toplam_kayit': toplam_kayit,
            'toplam_su_miktari': round(toplam_su, 2),
            'ortalama_su_miktari': round(ortalama_su, 2),
            'aktif_tesis_sayisi': aktif_tesis_sayisi
        })


class UrunKategorisiViewSet(viewsets.ModelViewSet):
    """
    Ürün kategorisi ViewSet
    """
    queryset = UrunKategorisi.objects.all()
    serializer_class = UrunKategorisiSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['isim']
    search_fields = ['isim', 'aciklama']
    ordering_fields = ['isim', 'olusturma_tarihi']
    ordering = ['isim']
    pagination_class = None  # Pagination'ı kaldır


class UrunViewSet(viewsets.ModelViewSet):
    """
    Ürün yönetimi ViewSet
    Tüm ürünler tüm sulama sistemlerinde görünür
    """
    queryset = Urun.objects.select_related('sulama__bolge').prefetch_related('kategori').all()
    serializer_class = UrunSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['sulama', 'kategori', 'isim']
    search_fields = ['isim', 'sulama__isim', 'kategori__isim']
    ordering_fields = ['isim', 'olusturma_tarihi']
    ordering = ['sulama', 'isim']
    pagination_class = None

    def get_queryset(self):
        """Tüm ürünleri getir - sulama bazlı filtreleme yok"""
        return super().get_queryset()

    def get_serializer_class(self):
        """Tüm görünümlerde UrunSerializer kullan"""
        return UrunSerializer

    @action(detail=False, methods=['get'])
    def ozet(self, request):
        """Ürün özet listesi"""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = UrunOzetSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def yillik_tuketimler(self, request, pk=None):
        """Ürünün yıllık tüketimleri"""
        urun = self.get_object()
        # Yeni model yapısında ürün detayları ayrı tabloda
        urun_detaylari = YillikUrunDetay.objects.filter(urun=urun).select_related('yillik_tuketim__sulama').order_by('-yillik_tuketim__yil')
        serializer = YillikUrunDetaySerializer(urun_detaylari, many=True)
        return Response(serializer.data)


class YillikGenelSuTuketimiViewSet(SulamaBazliMixin, viewsets.ModelViewSet):
    """
    Yıllık genel su tüketimi ViewSet
    """
    queryset = YillikGenelSuTuketimi.objects.select_related(
        'sulama__bolge'
    ).prefetch_related('urun_detaylari__urun').all()
    serializer_class = YillikGenelSuTuketimiSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['yil', 'sulama']
    search_fields = ['sulama__isim', 'urun_detaylari__urun__isim']
    ordering_fields = ['yil', 'olusturma_tarihi']
    ordering = ['-yil', 'sulama']
    pagination_class = None  # Pagination'ı kaldır

    def get_queryset(self):
        """Kullanıcının yetkili olduğu sulama sistemlerine ait verileri getir"""
        base_queryset = super().get_queryset()
        return self.filter_by_sulama_permission(base_queryset, 'sulama')

    @action(detail=False, methods=['get'])
    def yil_ozeti(self, request):
        """Yıl bazında özet istatistikler"""
        yil = request.query_params.get('yil')
        if not yil:
            return Response(
                {'error': 'Yıl parametresi gerekli'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            yil = int(yil)
        except ValueError:
            return Response(
                {'error': 'Geçersiz yıl değeri'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.get_queryset().filter(yil=yil)
        
        # Yeni model yapısı için aggregate'ler
        toplam_alan = 0
        toplam_tuketim = 0
        kayit_sayisi = queryset.count()
        
        # Her ana kayıt için detayları topla
        for ana_kayit in queryset:
            toplam_alan += ana_kayit.get_toplam_alan()
            toplam_tuketim += ana_kayit.get_toplam_su_tuketimi()
        
        ozet = {
            'toplam_alan': toplam_alan,
            'toplam_tuketim': toplam_tuketim,
            'ortalama_ciftlik_randi': queryset.aggregate(avg=Avg('ciftlik_randi'))['avg'],
            'ortalama_iletim_randi': queryset.aggregate(avg=Avg('iletim_randi'))['avg'],
            'kayit_sayisi': kayit_sayisi
        }
        
        # Sulama sistemi bazında grupla
        sulama_ozeti = []
        for ana_kayit in queryset.select_related('sulama__bolge').prefetch_related('urun_detaylari'):
            sulama_ozeti.append({
                'sulama__isim': ana_kayit.sulama.isim,
                'sulama__bolge__isim': ana_kayit.sulama.bolge.isim,
                'alan': ana_kayit.get_toplam_alan(),
                'tuketim': ana_kayit.get_toplam_su_tuketimi(),
                'urun_sayisi': ana_kayit.urun_detaylari.count()
            })
        
        return Response({
            'genel_ozet': ozet,
            'sulama_ozeti': list(sulama_ozeti)
        })

    @action(detail=False, methods=['get'])
    def karsilastirma(self, request):
        """Yıllar arası karşılaştırma"""
        yil1 = request.query_params.get('yil1')
        yil2 = request.query_params.get('yil2')
        
        if not yil1 or not yil2:
            return Response(
                {'error': 'İki yıl parametresi gerekli (yil1, yil2)'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            yil1, yil2 = int(yil1), int(yil2)
        except ValueError:
            return Response(
                {'error': 'Geçersiz yıl değerleri'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.get_queryset()
        
        # Yıl 1 verileri
        queryset1 = queryset.filter(yil=yil1)
        toplam_alan1 = sum(ana_kayit.get_toplam_alan() for ana_kayit in queryset1)
        toplam_tuketim1 = sum(ana_kayit.get_toplam_su_tuketimi() for ana_kayit in queryset1)
        
        veri1 = {
            'toplam_alan': toplam_alan1,
            'toplam_tuketim': toplam_tuketim1,
            'kayit_sayisi': queryset1.count()
        }
        
        # Yıl 2 verileri
        queryset2 = queryset.filter(yil=yil2)
        toplam_alan2 = sum(ana_kayit.get_toplam_alan() for ana_kayit in queryset2)
        toplam_tuketim2 = sum(ana_kayit.get_toplam_su_tuketimi() for ana_kayit in queryset2)
        
        veri2 = {
            'toplam_alan': toplam_alan2,
            'toplam_tuketim': toplam_tuketim2,
            'kayit_sayisi': queryset2.count()
        }
        
        # Değişim oranları
        degisim = {}
        for key in ['toplam_alan', 'toplam_tuketim']:
            if veri1[key] and veri2[key]:
                degisim[key] = round(((veri2[key] - veri1[key]) / veri1[key]) * 100, 2)
            else:
                degisim[key] = None
        
        return Response({
            f'{yil1}_veriler': veri1,
            f'{yil2}_veriler': veri2,
            'degisim_oranlari': degisim
        })

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Toplu su tüketimi verisi kaydetme"""
        data = request.data
        
        # Gelen veriyi validate et
        if not isinstance(data, dict):
            return Response(
                {'error': 'Geçersiz veri formatı'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        required_fields = ['sulama', 'yil', 'ciftlik_randi', 'iletim_randi', 'table_data']
        for field in required_fields:
            if field not in data:
                return Response(
                    {'error': f'{field} alanı gerekli'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        try:
            sulama_id = int(data['sulama'])
            yil = int(data['yil'])
            ciftlik_randi = float(data['ciftlik_randi'])
            iletim_randi = float(data['iletim_randi'])
            table_data = data['table_data']
            
            # Sulama sisteminin varlığını ve yetkisini kontrol et
            try:
                from .models import Sulama
                sulama = Sulama.objects.get(id=sulama_id)
                
                # Yetki kontrolü - SulamaBazliMixin'deki filter_by_sulama_permission kullan
                allowed_sulamalar = self.filter_by_sulama_permission(
                    Sulama.objects.all(), 'id'
                ).values_list('id', flat=True)
                
                if sulama_id not in allowed_sulamalar:
                    return Response(
                        {'error': 'Bu sulama sistemine erişim yetkiniz yok'}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
                    
            except Sulama.DoesNotExist:
                return Response(
                    {'error': 'Geçersiz sulama sistemi'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Önce mevcut yıl verilerini sil (ana tablo ve detaylar)
            self.get_queryset().filter(sulama=sulama, yil=yil).delete()
            
            # Ana kaydı oluştur
            yillik_tuketim = YillikGenelSuTuketimi.objects.create(
                yil=yil,
                sulama=sulama,
                ciftlik_randi=ciftlik_randi,
                iletim_randi=iletim_randi
            )
            
            # Ürün detaylarını hazırla
            urun_detaylari = []
            for row in table_data:
                if not row.get('urun') or not row.get('ekim_alani'):
                    continue  # Boş satırları atla
                
                try:
                    urun_id = int(row['urun'])
                    ekim_alani = float(row['ekim_alani'])
                    ekim_orani = float(row.get('ekim_orani', 100))
                    
                    # Su tüketimini hesapla (m³ cinsinden)
                    su_tuketimi = float(row.get('su_tuketimi', 0))
                    
                    urun_detay = YillikUrunDetay(
                        yillik_tuketim=yillik_tuketim,
                        urun_id=urun_id,
                        alan=ekim_alani,  # Artık hektar cinsinden
                        ekim_orani=ekim_orani,
                        su_tuketimi=su_tuketimi
                    )
                    urun_detaylari.append(urun_detay)
                    
                except (ValueError, TypeError) as e:
                    # Ana kayıt oluşturuldu ama hata var, ana kaydı da sil
                    yillik_tuketim.delete()
                    return Response(
                        {'error': f'Geçersiz veri: {str(e)}'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            if not urun_detaylari:
                # Ana kayıt oluşturuldu ama detay yok, ana kaydı da sil
                yillik_tuketim.delete()
                return Response(
                    {'error': 'Kaydedilecek geçerli veri bulunamadı'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Ürün detaylarını toplu kaydet
            YillikUrunDetay.objects.bulk_create(urun_detaylari)
            
            return Response({
                'message': f'{len(urun_detaylari)} adet ürün detayı başarıyla kaydedildi',
                'ana_kayit_id': yillik_tuketim.id,
                'urun_detay_sayisi': len(urun_detaylari),
                'yil': yil,
                'sulama': sulama.isim
            }, status=status.HTTP_201_CREATED)
            
        except (ValueError, TypeError) as e:
            return Response(
                {'error': f'Veri formatı hatası: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'error': f'Kaydetme hatası: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class YillikUrunDetayViewSet(SulamaBazliMixin, viewsets.ModelViewSet):
    """
    Yıllık ürün detay ViewSet
    """
    queryset = YillikUrunDetay.objects.select_related(
        'yillik_tuketim__sulama__bolge', 'urun'
    ).all()
    serializer_class = YillikUrunDetaySerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['yillik_tuketim__yil', 'yillik_tuketim__sulama', 'urun']
    search_fields = ['urun__isim', 'yillik_tuketim__sulama__isim']
    ordering_fields = ['yillik_tuketim__yil', 'alan', 'su_tuketimi', 'olusturma_tarihi']
    ordering = ['-yillik_tuketim__yil', 'urun__isim']

    def get_queryset(self):
        """Kullanıcının yetkili olduğu sulama sistemlerine ait verileri getir"""
        base_queryset = super().get_queryset()
        return self.filter_by_sulama_permission(base_queryset, 'yillik_tuketim__sulama')


class DashboardViewSet(SulamaBazliMixin, viewsets.ViewSet):
    """
    Dashboard verileri ViewSet
    Aylık su kullanım verilerinin karşılaştırmalı görünümü
    """
    permission_classes = [SulamaYetkisiPermission]

    @action(detail=False, methods=['get'])
    def aylik_su_kullanimi(self, request):
        """
        Aylık su kullanım karşılaştırması
        1. Şebekeye Alınan Su (GunlukSebekeyeAlinanSuMiktari)
        2. Depolama Tesisindeki Su (GunlukDepolamaTesisiSuMiktari) 
        3. Genel Su Tüketimi (YillikGenelSuTuketimi)
        """
        try:
            # Parametreler
            yil = request.GET.get('yil', datetime.now().year)
            sulama_id = request.GET.get('sulama')
            depolama_tesisi_id = request.GET.get('depolama_tesisi')
            
            try:
                yil = int(yil)
            except (ValueError, TypeError):
                yil = datetime.now().year

            # Base queryset'leri hazırla
            sebeke_qs = GunlukSebekeyeAlinanSuMiktari.objects.filter(tarih__year=yil)
            depolama_qs = GunlukDepolamaTesisiSuMiktari.objects.filter(tarih__year=yil)
            tuketim_qs = YillikGenelSuTuketimi.objects.filter(yil=yil)

            # Depolama tesisi filtreleme (öncelik)
            if depolama_tesisi_id:
                try:
                    depolama_tesisi_id = int(depolama_tesisi_id)
                    sebeke_qs = sebeke_qs.filter(kanal__depolama_tesisi_id=depolama_tesisi_id)
                    depolama_qs = depolama_qs.filter(depolama_tesisi_id=depolama_tesisi_id)
                    # Depolama tesisi seçilince o tesisin sulama sistemini al
                    try:
                        depolama_tesisi = DepolamaTesisi.objects.get(id=depolama_tesisi_id)
                        tuketim_qs = tuketim_qs.filter(sulama_id=depolama_tesisi.sulama_id)
                    except DepolamaTesisi.DoesNotExist:
                        pass
                except (ValueError, TypeError):
                    pass
            # Eğer depolama tesisi seçilmemişse sulama filtreleme
            elif sulama_id:
                try:
                    sulama_id = int(sulama_id)
                    sebeke_qs = sebeke_qs.filter(kanal__depolama_tesisi__sulama_id=sulama_id)
                    depolama_qs = depolama_qs.filter(depolama_tesisi__sulama_id=sulama_id)
                    tuketim_qs = tuketim_qs.filter(sulama_id=sulama_id)
                except (ValueError, TypeError):
                    pass

            # Yetki kontrolü
            sebeke_qs = self.filter_by_sulama_permission(sebeke_qs, 'kanal__depolama_tesisi__sulama')
            depolama_qs = self.filter_by_sulama_permission(depolama_qs, 'depolama_tesisi__sulama')
            tuketim_qs = self.filter_by_sulama_permission(tuketim_qs, 'sulama')

            # Aylar için boş veri yapısı
            aylik_veri = {}
            aylar = [
                ('01', 'Ocak'), ('02', 'Şubat'), ('03', 'Mart'), ('04', 'Nisan'),
                ('05', 'Mayıs'), ('06', 'Haziran'), ('07', 'Temmuz'), ('08', 'Ağustos'),
                ('09', 'Eylül'), ('10', 'Ekim'), ('11', 'Kasım'), ('12', 'Aralık')
            ]

            for ay_no, ay_isim in aylar:
                aylik_veri[ay_no] = {
                    'ay': ay_isim,
                    'ay_no': int(ay_no),
                    'sebeke_su': 0,  # Şebekeye alınan toplam su (m³)
                    'depolama_su': 0,  # Depolamadaki ortalama su (m³)
                    'tuketim_su': 0,  # Genel tüketim (m³)
                    'sebeke_kayit_sayisi': 0,
                    'depolama_kayit_sayisi': 0,
                    'tuketim_kayit_sayisi': 0
                }

            # 1. Şebekeye Alınan Su (Aylık toplam)
            sebeke_veriler = sebeke_qs.annotate(
                ay=Extract('tarih', 'month')
            ).values('ay').annotate(
                toplam_su=Sum('su_miktari'),
                kayit_sayisi=Count('id')
            )

            for veri in sebeke_veriler:
                ay = f"{veri['ay']:02d}"  # 01, 02, 03 formatında
                if ay in aylik_veri:
                    aylik_veri[ay]['sebeke_su'] = float(veri['toplam_su'] or 0)
                    aylik_veri[ay]['sebeke_kayit_sayisi'] = veri['kayit_sayisi']

            # 2. Depolama Tesisindeki Su (Her ay için son tarihteki toplam)
            from django.db.models import Max
            
            for ay_idx, (ay_no, _) in enumerate(aylar):
                ay_num = int(ay_no)
                
                # Bu ay için her depolama tesisinin son kaydını al
                aylik_depolama_toplam = 0
                aylik_depolama_kayit = 0
                
                # Yetkili depolama tesislerini al
                yetkili_tesisler = DepolamaTesisi.objects.filter(
                    gunluk_depolama_tesisi_su_miktarlari__in=depolama_qs
                ).distinct()
                
                for tesis in yetkili_tesisler:
                    # Bu tesis için bu ayın son kaydını al
                    son_kayit = depolama_qs.filter(
                        depolama_tesisi=tesis,
                        tarih__month=ay_num
                    ).order_by('-tarih').first()
                    
                    if son_kayit:
                        aylik_depolama_toplam += float(son_kayit.su_miktari or 0)
                        aylik_depolama_kayit += 1
                
                aylik_veri[ay_no]['depolama_su'] = aylik_depolama_toplam
                aylik_veri[ay_no]['depolama_kayit_sayisi'] = aylik_depolama_kayit

            # 3. Genel Su Tüketimi (Yeni model yapısı ile hesaplama)
            ana_tuketim_kayitlari = tuketim_qs.prefetch_related('urun_detaylari__urun').all()
            tuketim_kayit_sayisi = len(ana_tuketim_kayitlari)
            
            # Her ay için tüketim hesapla
            for ay_idx, (ay_no, _) in enumerate(aylar):
                aylik_tuketim = 0
                
                for ana_kayit in ana_tuketim_kayitlari:
                    # Bu ana kayıt altındaki tüm ürün detayları
                    for urun_detay in ana_kayit.urun_detaylari.all():
                        # Alan (hektar)
                        alan_ha = urun_detay.alan
                        
                        # Ürünün aylık UR değerlerini al
                        urun = urun_detay.urun
                        aylik_ur_degerleri = [
                            urun.ocak or 0, urun.subat or 0, urun.mart or 0,
                            urun.nisan or 0, urun.mayis or 0, urun.haziran or 0,
                            urun.temmuz or 0, urun.agustos or 0, urun.eylul or 0,
                            urun.ekim or 0, urun.kasim or 0, urun.aralik or 0
                        ]
                        
                        if len(aylik_ur_degerleri) >= 12:
                            # Bu ayın UR değeri
                            bu_ay_ur = aylik_ur_degerleri[ay_idx]
                            
                            # NET SU İHTİYACI hesaplaması (Frontend'deki gibi)
                            # Alan × UR değeri ÷ 100000 (hm³ cinsinden)
                            net_su_degeri = (alan_ha * bu_ay_ur) / 100000
                            
                            # m³ cinsine çevir (× 1000000)
                            bu_ay_tuketim_m3 = net_su_degeri * 1000000
                            aylik_tuketim += bu_ay_tuketim_m3
                        else:
                            # UR değerleri yoksa eşit dağıt (fallback)
                            aylik_tuketim += urun_detay.su_tuketimi / 12
                
                aylik_veri[ay_no]['tuketim_su'] = float(aylik_tuketim)
                aylik_veri[ay_no]['tuketim_kayit_sayisi'] = tuketim_kayit_sayisi

            # Toplam yıllık tüketimi hesapla
            toplam_yillik_tuketim = sum(aylik_veri[ay_no]['tuketim_su'] for ay_no, _ in aylar)

            # Sıralı liste haline getir
            aylik_liste = [aylik_veri[ay_no] for ay_no, _ in aylar]

            # Güncel depo durumu vs Gelecek ihtiyaç analizi
            guncel_ay = datetime.now().month
            guncel_depo_miktari = 0
            gelecek_ihtiyac = 0
            
            # Güncel ayın depo miktarı (tüm depoların toplamı)
            if guncel_ay <= 12:
                guncel_ay_str = f"{guncel_ay:02d}"
                guncel_depo_miktari = aylik_veri[guncel_ay_str]['depolama_su']
            
            # Güncel aydan sonraki ayların planlanan tüketimi
            for ay_idx in range(guncel_ay, 12):  # Güncel aydan sonraki aylar
                ay_no = f"{ay_idx + 1:02d}"
                if ay_no in aylik_veri:
                    gelecek_ihtiyac += aylik_veri[ay_no]['tuketim_su']
            
            # Yeterlilik durumu
            yeterlilik_durumu = "Yeterli" if guncel_depo_miktari >= gelecek_ihtiyac else "Yetersiz"
            yeterlilik_orani = (guncel_depo_miktari / gelecek_ihtiyac * 100) if gelecek_ihtiyac > 0 else 100

            # Toplam istatistikler
            toplam_istatistikler = {
                'toplam_sebeke_su': sum(veri['sebeke_su'] for veri in aylik_liste),
                'toplam_depolama_su': sum(veri['depolama_su'] for veri in aylik_liste),
                'toplam_tuketim_su': toplam_yillik_tuketim,
                'toplam_sebeke_kayit': sum(veri['sebeke_kayit_sayisi'] for veri in aylik_liste),
                'toplam_depolama_kayit': sum(veri['depolama_kayit_sayisi'] for veri in aylik_liste),
                'toplam_tuketim_kayit': tuketim_kayit_sayisi,
                'yil': yil,
                'sulama_id': sulama_id,
                'depolama_tesisi_id': depolama_tesisi_id,
                # Yeni eklenen analizler
                'guncel_depo_miktari': guncel_depo_miktari,
                'gelecek_ihtiyac': gelecek_ihtiyac,
                'yeterlilik_durumu': yeterlilik_durumu,
                'yeterlilik_orani': round(yeterlilik_orani, 1),
                'guncel_ay': guncel_ay
            }

            return Response({
                'aylik_veriler': aylik_liste,
                'istatistikler': toplam_istatistikler,
                'success': True
            })

        except Exception as e:
            return Response(
                {'error': f'Dashboard verileri alınırken hata oluştu: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@csrf_exempt
def export_to_excel_with_template(request):
    if request.method == "POST":
        data = json.loads(request.body)
        formData = data.get("formData", {})
        tableData = data.get("tableData", [])
        results = data.get("results", {})

        yil = str(formData.get("yil", "2024"))
        sulama_adi = formData.get("sulama", "SulamaAdı")
        kurum_adi = formData.get("kurumAdi", "Kurum")
        genel_adi = "GenelSulamaPlanlaması"

        # Şablon dosyanın tam yolu:
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_path = os.path.join(BASE_DIR, "excel_templates", "Kitap1.xlsx")
        temp_path = os.path.join(BASE_DIR, "excel_templates", f"temp_{yil}_{sulama_adi}.xlsx")

        shutil.copy(template_path, temp_path)
        wb = load_workbook(temp_path)
        ws = wb.active

        ws["A1"] = f"{yil} {kurum_adi} {sulama_adi} {genel_adi}"

        start_row = 4
        for i, row in enumerate(tableData, start=start_row):
            ws[f"A{i}"] = row.get("urun", "")
            ws[f"B{i}"] = row.get("ekim_alani", "")
            ws[f"C{i}"] = row.get("ekim_orani", "")
            ur_values = row.get("ur_values", [])
            for j, val in enumerate(ur_values):
                col_letter = chr(ord("D") + j)
                ws[f"{col_letter}{i}"] = val
            ws[f"P{i}"] = row.get("toplam_ur", "")
            ws[f"Q{i}"] = row.get("su_tuketimi", "")

        result_labels = [
            ("NET SU İHTİYACI (hm³)", results.get("net_su_aylik", []), results.get("net_su_toplam", "")),
            ("ÇİFTLİK SU İHTİYACI (hm³)", results.get("ciftlik_su_aylik", []), results.get("ciftlik_su_toplam", "")),
            ("BRÜT SU İHTİYACI (hm³)", results.get("brut_su_aylik", []), results.get("brut_su_toplam", "")),
        ]
        for offset, (label, aylik, toplam) in enumerate(result_labels):
            result_row = start_row + len(tableData) + 2 + offset
            ws[f"A{result_row}"] = label
            for j, val in enumerate(aylik):
                col_letter = chr(ord("D") + j)
                ws[f"{col_letter}{result_row}"] = val
            ws[f"P{result_row}"] = toplam

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        wb.close()
        os.remove(temp_path)
        output.seek(0)

        filename = f"{yil}_{sulama_adi}_{genel_adi}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    return JsonResponse({"error": "Invalid request"}, status=400)


class MakinaViewSet(SulamaBazliMixin, viewsets.ModelViewSet):
    """
    Makina yönetimi ViewSet
    """
    queryset = Makina.objects.select_related('sulama__bolge').all()
    serializer_class = MakinaSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['sulama', 'makina_tipi', 'durum']
    search_fields = ['isim', 'plaka', 'model', 'sulama__isim']
    ordering_fields = ['isim', 'makina_tipi', 'durum', 'olusturma_tarihi']
    ordering = ['sulama__bolge__isim', 'sulama__isim', 'isim']
    pagination_class = None

    def get_queryset(self):
        """Kullanıcının yetkili olduğu sulama sistemlerine ait makineleri getir"""
        base_queryset = super().get_queryset()
        return self.filter_by_sulama_permission(base_queryset, 'sulama')

    @action(detail=False, methods=['get'])
    def harita_verileri(self, request):
        """Harita için makina konum verilerini getir"""
        try:
            makinalar = self.get_queryset().prefetch_related('konumlar', 'isler')
            harita_verileri = []
            
            for makina in makinalar:
                son_konum = makina.konumlar.first()
                aktif_is = makina.isler.filter(durum='devam_ediyor').first()
                
                # Varsayılan konum (Samsun merkez)
                varsayilan_enlem = 41.2867
                varsayilan_boylam = 36.3300
                
                veri = {
                    'id': makina.id,
                    'birlik_no': makina.birlik_no,
                    'isim': makina.isim,
                    'makina_tipi': makina.makina_tipi,
                    'makina_tipi_display': makina.get_makina_tipi_display(),
                    'plaka': makina.plaka,
                    'model': makina.model,
                    'yil': makina.yil,
                    'durum': makina.durum,
                    'durum_display': makina.get_durum_display(),
                    'enlem': float(son_konum.enlem) if son_konum else varsayilan_enlem,
                    'boylam': float(son_konum.boylam) if son_konum else varsayilan_boylam,
                    'kayit_zamani': son_konum.kayit_zamani if son_konum else None,
                    'konum_var': son_konum is not None,
                    'aktif_is': None
                }
                
                if aktif_is:
                    veri['aktif_is'] = {
                        'id': aktif_is.id,
                        'is_tipi': aktif_is.is_tipi,
                        'is_tipi_display': aktif_is.get_is_tipi_display(),
                        'baslik': aktif_is.baslik,
                        'calistigi_yer': aktif_is.calistigi_yer,
                        'baslangic_zamani': aktif_is.baslangic_zamani,
                        'is_enlem': float(aktif_is.enlem) if aktif_is.enlem else None,
                        'is_boylam': float(aktif_is.boylam) if aktif_is.boylam else None
                    }
                
                harita_verileri.append(veri)
            
            return Response({
                'makinalar': harita_verileri,
                'success': True
            })
            
        except Exception as e:
            return Response(
                {'error': f'Harita verileri alınırken hata oluştu: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class MakinaKonumViewSet(SulamaBazliMixin, viewsets.ModelViewSet):
    """
    Makina konum takibi ViewSet
    """
    queryset = MakinaKonum.objects.select_related('makina__sulama__bolge').all()
    serializer_class = MakinaKonumSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['makina', 'makina__sulama', 'motor_calisma']
    search_fields = ['makina__isim', 'makina__plaka']
    ordering_fields = ['kayit_zamani', 'makina__isim']
    ordering = ['-kayit_zamani']

    def get_queryset(self):
        """Kullanıcının yetkili olduğu sulama sistemlerine ait konum verilerini getir"""
        base_queryset = super().get_queryset()
        return self.filter_by_sulama_permission(base_queryset, 'makina__sulama')

    @action(detail=False, methods=['post'])
    def toplu_konum_guncelle(self, request):
        """Toplu konum güncelleme (IoT cihazlarından)"""
        try:
            konum_verileri = request.data.get('konumlar', [])
            guncellenen_sayisi = 0
            
            for veri in konum_verileri:
                makina_id = veri.get('makina_id')
                enlem = veri.get('enlem')
                boylam = veri.get('boylam')
                hiz = veri.get('hiz')
                yon = veri.get('yon')
                yakit_seviyesi = veri.get('yakit_seviyesi')
                motor_calisma = veri.get('motor_calisma', True)
                
                if makina_id and enlem and boylam:
                    MakinaKonum.objects.create(
                        makina_id=makina_id,
                        enlem=enlem,
                        boylam=boylam
                    )
                    guncellenen_sayisi += 1
            
            return Response({
                'message': f'{guncellenen_sayisi} makina konumu güncellendi',
                'guncellenen_sayisi': guncellenen_sayisi,
                'success': True
            })
            
        except Exception as e:
            return Response(
                {'error': f'Konum güncellenirken hata oluştu: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class MakinaIsViewSet(SulamaBazliMixin, viewsets.ModelViewSet):
    """
    Makina iş takibi ViewSet
    """
    queryset = MakinaIs.objects.select_related('makina__sulama__bolge').all()
    serializer_class = MakinaIsSerializer
    permission_classes = [SulamaYetkisiPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['makina', 'makina__sulama', 'is_tipi', 'durum']
    search_fields = ['baslik', 'aciklama', 'makina__isim']
    ordering_fields = ['baslangic_zamani', 'bitis_zamani', 'durum']
    ordering = ['-baslangic_zamani']
    pagination_class = None  # Pagination'ı kaldır

    def get_queryset(self):
        """Kullanıcının yetkili olduğu sulama sistemlerine ait iş verilerini getir"""
        base_queryset = super().get_queryset()
        return self.filter_by_sulama_permission(base_queryset, 'makina__sulama')

    @action(detail=True, methods=['post'])
    def is_baslat(self, request, pk=None):
        """İşi başlat"""
        try:
            makina_is = self.get_object()
            makina_is.durum = 'devam_ediyor'
            makina_is.save()
            
            return Response({
                'message': 'İş başlatıldı',
                'success': True
            })
            
        except Exception as e:
            return Response(
                {'error': f'İş başlatılırken hata oluştu: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def is_tamamla(self, request, pk=None):
        """İşi tamamla"""
        try:
            makina_is = self.get_object()
            makina_is.durum = 'tamamlandi'
            makina_is.bitis_zamani = timezone.now()
            makina_is.save()
            
            return Response({
                'message': 'İş tamamlandı',
                'success': True
            })
            
        except Exception as e:
            return Response(
                {'error': f'İş tamamlanırken hata oluştu: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )